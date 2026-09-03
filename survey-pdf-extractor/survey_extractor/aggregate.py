"""抽出結果 JSON を pandas で結合し、Excel 3シートを出力する。"""

from __future__ import annotations

import statistics
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Config, Question
from .normalize import display_value

SHEET_RAW = "生データ"
SHEET_SUMMARY = "集計"
SHEET_REVIEW = "要確認"

REVIEW_CONFIDENCE = ("low", "medium")
_CONFIDENCE_SORT = {"low": 0, "medium": 1, "high": 2}

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_LOW_FILL = PatternFill("solid", fgColor="F8CBAD")   # 低信頼度
_MED_FILL = PatternFill("solid", fgColor="FFF2CC")   # 中信頼度
_MAX_COL_WIDTH = 60


# ---------------------------------------------------------------------------
# シート組み立て
# ---------------------------------------------------------------------------
def _answer(record: dict[str, Any], qid: str) -> dict[str, Any]:
    return record.get("answers", {}).get(qid) or {
        "value": None,
        "raw_text": None,
        "status": "unreadable",
        "confidence": "low",
        "page": None,
        "flags": ["missing_in_record"],
    }


def build_raw_sheet(config: Config, records: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for record in records:
        row: dict[str, Any] = {
            "回答者ID": record.get("respondent_id", ""),
            "回答者名": record.get("respondent_name") or "",
            "元ファイル": record.get("source_file", ""),
        }
        for q in config.questions:
            row[q.label] = display_value(q, _answer(record, q.id))
        rows.append(row)

    columns = ["回答者ID", "回答者名", "元ファイル", *(q.label for q in config.questions)]
    return pd.DataFrame(rows, columns=columns)


def _counts_for_choice(
    q: Question, records: list[dict[str, Any]]
) -> tuple[dict[str, int], int, int, int, int]:
    """(選択肢別件数, 有効回答数, 未記入, 判読不能, 選択肢外)"""
    counts = {c: 0 for c in q.choices}
    answered = blank = unreadable = other = 0

    for record in records:
        a = _answer(record, q.id)
        value = a.get("value")
        status = a.get("status")

        if q.type == "single_choice":
            selected = [value] if isinstance(value, str) else []
        else:
            selected = [v for v in (value or []) if isinstance(v, str)]

        known = [v for v in selected if v in counts]
        unknown = [v for v in selected if v not in counts]

        if known or unknown:
            answered += 1
            for v in known:
                counts[v] += 1
            if unknown:
                other += 1
        elif status == "blank":
            blank += 1
        else:
            unreadable += 1

    return counts, answered, blank, unreadable, other


def build_summary_sheet(config: Config, records: list[dict[str, Any]]) -> pd.DataFrame:
    total = len(records)
    rows: list[dict[str, Any]] = []

    def add(q: Question, item: str, value: Any, ratio: Any = None) -> None:
        rows.append(
            {
                "設問ID": q.id,
                "設問文": q.text,
                "種別": q.type,
                "項目": item,
                "値": value,
                "構成比": ratio,
            }
        )

    for q in config.questions:
        if q.is_choice:
            counts, answered, blank, unreadable, other = _counts_for_choice(q, records)
            denominator = answered or 1
            label = "有効回答数" if q.type == "single_choice" else "有効回答者数"
            add(q, label, answered, answered / total if total else None)
            for choice in q.choices:
                add(q, choice, counts[choice], counts[choice] / denominator)
            if other:
                add(q, "選択肢以外の記入（要確認）", other, other / denominator)
            add(q, "未記入", blank, blank / total if total else None)
            add(q, "判読不能", unreadable, unreadable / total if total else None)

        elif q.type == "number":
            values: list[float] = []
            blank = unreadable = 0
            for record in records:
                a = _answer(record, q.id)
                v = a.get("value")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    values.append(float(v))
                elif a.get("status") == "blank":
                    blank += 1
                else:
                    unreadable += 1
            add(q, "件数", len(values), len(values) / total if total else None)
            if values:
                add(q, "平均", round(statistics.mean(values), 2))
                add(q, "中央値", round(statistics.median(values), 2))
                add(q, "最小", min(values))
                add(q, "最大", max(values))
            add(q, "未記入", blank, blank / total if total else None)
            add(q, "判読不能", unreadable, unreadable / total if total else None)

        else:  # free_text
            written = blank = unreadable = 0
            for record in records:
                a = _answer(record, q.id)
                if a.get("value"):
                    written += 1
                elif a.get("status") == "blank":
                    blank += 1
                else:
                    unreadable += 1
            add(q, "記入あり", written, written / total if total else None)
            add(q, "未記入", blank, blank / total if total else None)
            add(q, "判読不能", unreadable, unreadable / total if total else None)

    return pd.DataFrame(
        rows, columns=["設問ID", "設問文", "種別", "項目", "値", "構成比"]
    )


def build_review_sheet(config: Config, records: list[dict[str, Any]]) -> pd.DataFrame:
    order = {q.id: i for i, q in enumerate(config.questions)}
    rows = []
    for record in records:
        for q in config.questions:
            a = _answer(record, q.id)
            confidence = a.get("confidence")
            flags = a.get("flags") or []
            if confidence not in REVIEW_CONFIDENCE and not flags:
                continue
            page = a.get("page")
            rows.append(
                {
                    "回答者ID": record.get("respondent_id", ""),
                    "設問ID": q.id,
                    "設問文": q.text,
                    "抽出値": display_value(q, a),
                    "読み取り原文": a.get("raw_text") or "",
                    "状態": a.get("status", ""),
                    "confidence": confidence,
                    "該当ページ": page if page is not None else "",
                    "備考": " / ".join(flags),
                    "_sort": (_CONFIDENCE_SORT.get(confidence, 0), record.get("respondent_id", ""), order[q.id]),
                }
            )

    columns = [
        "回答者ID",
        "設問ID",
        "設問文",
        "抽出値",
        "読み取り原文",
        "状態",
        "confidence",
        "該当ページ",
        "備考",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)

    rows.sort(key=lambda r: r["_sort"])
    for r in rows:
        del r["_sort"]
    return pd.DataFrame(rows, columns=columns)


# ---------------------------------------------------------------------------
# Excel 出力
# ---------------------------------------------------------------------------
def _display_width(value: Any) -> int:
    text = "" if value is None else str(value)
    width = 0
    for ch in text.splitlines()[0] if text else "":
        width += 2 if unicodedata.east_asian_width(ch) in "WF" else 1
    return width


def _style_sheet(ws, df: pd.DataFrame, wrap_columns: set[str] = frozenset()) -> None:
    if ws.max_row == 0:
        return
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "B2"
    if len(df.columns):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{max(ws.max_row, 1)}"

    for idx, column in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        widest = max(
            [_display_width(column)] + [_display_width(v) for v in df[column].tolist()[:200]]
        )
        ws.column_dimensions[letter].width = min(max(widest + 2, 10), _MAX_COL_WIDTH)
        if column in wrap_columns:
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _highlight_raw_confidence(ws, config: Config, records: list[dict[str, Any]]) -> None:
    """生データシートで、低・中信頼度のセルに色を付ける。"""
    first_question_col = 4  # A:回答者ID B:回答者名 C:元ファイル
    for row_idx, record in enumerate(records, start=2):
        for col_offset, q in enumerate(config.questions):
            confidence = _answer(record, q.id).get("confidence")
            if confidence == "low":
                ws.cell(row=row_idx, column=first_question_col + col_offset).fill = _LOW_FILL
            elif confidence == "medium":
                ws.cell(row=row_idx, column=first_question_col + col_offset).fill = _MED_FILL


def write_excel(
    config: Config, records: list[dict[str, Any]], output_path: Path
) -> tuple[Path, dict[str, int]]:
    """3シートの Excel を書き出し、(パス, 概要カウント) を返す。"""
    raw_df = build_raw_sheet(config, records)
    summary_df = build_summary_sheet(config, records)
    review_df = build_review_sheet(config, records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name=SHEET_RAW, index=False)
        summary_df.to_excel(writer, sheet_name=SHEET_SUMMARY, index=False)
        review_df.to_excel(writer, sheet_name=SHEET_REVIEW, index=False)

        book = writer.book
        _style_sheet(book[SHEET_RAW], raw_df, wrap_columns=set(raw_df.columns[3:]))
        _style_sheet(book[SHEET_SUMMARY], summary_df)
        _style_sheet(book[SHEET_REVIEW], review_df, wrap_columns={"設問文", "抽出値", "読み取り原文"})

        _highlight_raw_confidence(book[SHEET_RAW], config, records)

        ratio_col = get_column_letter(list(summary_df.columns).index("構成比") + 1)
        for cell in book[SHEET_SUMMARY][ratio_col][1:]:
            cell.number_format = "0.0%"

    return output_path, {
        "respondents": len(raw_df),
        "summary_rows": len(summary_df),
        "review_rows": len(review_df),
    }
